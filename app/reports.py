import io
from datetime import date
from collections import defaultdict
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm


def format_amount(amount: float, currency: str) -> str:
    if currency == "INR":
        return f"\u20b9{amount:,.0f}"
    if currency == "USD":
        return f"${amount:,.2f}"
    if currency == "EUR":
        return f"\u20ac{amount:,.2f}"
    return f"{currency} {amount:,.2f}"


def format_amount_pdf(amount: float, currency: str) -> str:
    """ASCII-safe version for PDF (Helvetica doesn't support ₹ or €)."""
    if currency == "INR":
        return f"Rs.{amount:,.0f}"
    if currency == "USD":
        return f"${amount:,.2f}"
    if currency == "EUR":
        return f"EUR {amount:,.2f}"
    return f"{currency} {amount:,.2f}"


def generate_pdf_report(expenses: list[dict], user: dict, month: date) -> bytes:
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2 * cm, leftMargin=2 * cm,
        topMargin=2 * cm, bottomMargin=2 * cm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "ReportTitle", parent=styles["Heading1"],
        fontSize=18, spaceAfter=4, textColor=colors.HexColor("#1a1a2e")
    )
    subtitle_style = ParagraphStyle(
        "Subtitle", parent=styles["Normal"],
        fontSize=9, textColor=colors.grey, spaceAfter=18
    )
    section_style = ParagraphStyle(
        "Section", parent=styles["Heading2"],
        fontSize=11, spaceBefore=14, spaceAfter=6,
        textColor=colors.HexColor("#1a1a2e")
    )

    elements = []

    month_str = month.strftime("%B %Y")
    elements.append(Paragraph(f"Expense Report \u2014 {month_str}", title_style))
    elements.append(Paragraph(f"Account: {user['phone_number']}", subtitle_style))

    if not expenses:
        elements.append(Paragraph("No expenses recorded for this period.", styles["Normal"]))
        doc.build(elements)
        return buffer.getvalue()

    show_inr_equiv = user.get("default_currency", "INR") in ("USD", "EUR")

    # Expenses table
    elements.append(Paragraph("All Expenses", section_style))

    if show_inr_equiv:
        header = ["Date", "Description", "Category", "Amount", "INR Equiv."]
        col_widths = [2.0 * cm, 7.5 * cm, 3.5 * cm, 2.8 * cm, 2.2 * cm]
    else:
        header = ["Date", "Description", "Category", "Amount"]
        col_widths = [2.2 * cm, 9 * cm, 4 * cm, 2.8 * cm]

    rows = [header]
    for e in expenses:
        d = date.fromisoformat(e["expense_date"])
        row = [
            d.strftime("%d %b"),
            e["description"][:34],
            e["category"],
            format_amount_pdf(e["amount"], e["currency"]),
        ]
        if show_inr_equiv:
            inr_eq = e.get("inr_equivalent")
            row.append(format_amount_pdf(inr_eq, "INR") if inr_eq else "-")
        rows.append(row)

    amount_col = 3
    table = Table(rows, colWidths=col_widths)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dee2e6")),
        ("ALIGN", (amount_col, 0), (-1, -1), "RIGHT"),
        ("PADDING", (0, 0), (-1, -1), 5),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 0.4 * cm))

    # Category summary
    elements.append(Paragraph("Summary by Category", section_style))

    by_category: dict[str, dict[str, float]] = defaultdict(lambda: defaultdict(float))
    for e in expenses:
        by_category[e["category"]][e["currency"]] += e["amount"]

    summary_rows = [["Category", "Total"]]
    for category, currencies in sorted(by_category.items()):
        parts = [format_amount_pdf(amt, cur) for cur, amt in currencies.items()]
        summary_rows.append([category, "  +  ".join(parts)])

    summary_table = Table(summary_rows, colWidths=[8 * cm, 5 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1a1a2e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f4f6f9")]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#dee2e6")),
        ("ALIGN", (1, 0), (1, -1), "RIGHT"),
        ("PADDING", (0, 0), (-1, -1), 5),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 0.3 * cm))

    # Grand total
    totals: dict[str, float] = defaultdict(float)
    for e in expenses:
        totals[e["currency"]] += e["amount"]
    total_str = "  +  ".join(format_amount_pdf(amt, cur) for cur, amt in totals.items())
    grand_total_line = f"<b>Grand Total: {total_str}</b>"
    if show_inr_equiv:
        total_inr = sum(e.get("inr_equivalent") or 0 for e in expenses if e.get("inr_equivalent"))
        if total_inr:
            grand_total_line += f"  <i>(Rs.{total_inr:,.0f} total INR equiv.)</i>"
    elements.append(Paragraph(grand_total_line, styles["Normal"]))
    elements.append(Spacer(1, 0.2 * cm))
    elements.append(Paragraph(
        f"<i>{len(expenses)} transaction{'s' if len(expenses) != 1 else ''} \u2014 Generated by Kanak</i>",
        ParagraphStyle("Footer", parent=styles["Normal"], fontSize=7, textColor=colors.grey)
    ))

    doc.build(elements)
    return buffer.getvalue()


def generate_excel_report(expenses: list[dict], month: date) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = month.strftime("%B %Y")

    header_fill = PatternFill(fill_type="solid", fgColor="1a1a2e")
    header_font = Font(bold=True, color="FFFFFF", size=10)
    normal_font = Font(size=10)
    alt_fill = PatternFill(fill_type="solid", fgColor="F4F6F9")

    headers = ["Date", "Description", "Category", "Amount", "Currency", "INR Equivalent"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

    for row_idx, e in enumerate(expenses, 2):
        fill = alt_fill if row_idx % 2 == 0 else PatternFill()
        values = [
            e["expense_date"],
            e["description"],
            e["category"],
            e["amount"],
            e["currency"],
            e.get("inr_equivalent") or "",
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.font = normal_font
            cell.fill = fill

    # Right-align amount columns
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=6):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    # Auto-width columns
    col_widths = [12, 36, 18, 12, 10, 16]
    for col, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Summary sheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Category").font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value="Total").font = header_font
    ws2.cell(row=1, column=2).fill = header_fill

    by_category: dict[str, dict[str, float]] = {}
    for e in expenses:
        cat = e["category"]
        cur = e["currency"]
        by_category.setdefault(cat, {})
        by_category[cat][cur] = by_category[cat].get(cur, 0) + e["amount"]

    for row_idx, (cat, currencies) in enumerate(sorted(by_category.items()), 2):
        total_str = " + ".join(f"{format_amount(amt, cur)}" for cur, amt in currencies.items())
        ws2.cell(row=row_idx, column=1, value=cat).font = normal_font
        ws2.cell(row=row_idx, column=2, value=total_str).font = normal_font

    ws2.column_dimensions["A"].width = 20
    ws2.column_dimensions["B"].width = 24

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
